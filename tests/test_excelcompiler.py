# -*- coding: UTF-8 -*-
#
# Copyright 2011-2019 by Dirk Gorissen, Stephen Rauch and Contributors
# All rights reserved.
# This file is part of the Pycel Library, Licensed under GPLv3 (the 'License')
# You may not use this work except in compliance with the License.
# You may obtain a copy of the Licence at:
#   https://www.gnu.org/licenses/gpl-3.0.en.html

import json
import os
import shutil
from unittest import mock

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from ruamel.yaml import YAML

from pycel.excelcompiler import _Cell, _CellRange, ExcelCompiler
from pycel.excelformula import FormulaParserError, UnknownFunction
from pycel.excelutil import (
    AddressCell,
    AddressRange,
    flatten,
    list_like,
    NA_ERROR,
    NULL_ERROR,
)
from pycel.excelwrapper import ExcelWrapper


# ::TODO:: need some rectangular ranges for testing


def test_end_2_end(excel, fixture_xls_path):
    # load & compile the file to a graph, starting from D1
    for excel_compiler in (ExcelCompiler(excel=excel),
                           ExcelCompiler(excel=excel.workbook),
                           ExcelCompiler(fixture_xls_path)):

        # test evaluation
        assert -0.02286 == round(excel_compiler.evaluate('Sheet1!D1'), 5)

        excel_compiler.set_value('Sheet1!A1', 200)
        assert -0.00331 == round(excel_compiler.evaluate('Sheet1!D1'), 5)


def test_no_sheet_given(excel_compiler):
    sh1_value = excel_compiler.evaluate('Sheet1!A1')

    excel_compiler.excel.set_sheet('Sheet1')
    value = excel_compiler.evaluate('A1')
    assert sh1_value == value

    excel_compiler.excel.set_sheet('Sheet2')
    value = excel_compiler.evaluate('A1')
    assert sh1_value != value


def test_round_trip_through_json_yaml_and_pickle(
        excel_compiler, fixture_xls_path):
    excel_compiler.evaluate('Sheet1!D1')
    excel_compiler.extra_data = {1: 3}
    excel_compiler.to_file(file_types=('pickle', ))
    excel_compiler.to_file(file_types=('yml', ))
    excel_compiler.to_file(file_types=('json', ))

    # read the spreadsheet from json, yaml and pickle
    excel_compiler_json = ExcelCompiler.from_file(
        excel_compiler.filename + '.json')
    excel_compiler_yaml = ExcelCompiler.from_file(
        excel_compiler.filename + '.yml')
    excel_compiler = ExcelCompiler.from_file(excel_compiler.filename)

    # test evaluation
    assert -0.02286 == round(excel_compiler_json.evaluate('Sheet1!D1'), 5)
    assert -0.02286 == round(excel_compiler_yaml.evaluate('Sheet1!D1'), 5)
    assert -0.02286 == round(excel_compiler.evaluate('Sheet1!D1'), 5)

    excel_compiler_json.set_value('Sheet1!A1', 200)
    assert -0.00331 == round(excel_compiler_json.evaluate('Sheet1!D1'), 5)

    excel_compiler_yaml.set_value('Sheet1!A1', 200)
    assert -0.00331 == round(excel_compiler_yaml.evaluate('Sheet1!D1'), 5)

    excel_compiler.set_value('Sheet1!A1', 200)
    assert -0.00331 == round(excel_compiler.evaluate('Sheet1!D1'), 5)


def test_filename_ext(excel_compiler, fixture_xls_path):
    excel_compiler.evaluate('Sheet1!D1')
    excel_compiler.extra_data = {1: 3}
    pickle_name = excel_compiler.filename + '.pkl'
    yaml_name = excel_compiler.filename + '.yml'
    json_name = excel_compiler.filename + '.json'

    for name in (pickle_name, yaml_name, json_name):
        try:
            os.unlink(name)
        except FileNotFoundError:
            pass

    excel_compiler.to_file(excel_compiler.filename)
    excel_compiler.to_file(json_name, file_types=('json', ))

    assert os.path.exists(pickle_name)
    assert os.path.exists(yaml_name)
    assert os.path.exists(json_name)


def test_deserialize_filename(
        excel_compiler, fixture_xls_path, serialization_override_path):

    for serialization_filename, expected in (
        # When the serialization path is different than the workbook path
        (serialization_override_path, excel_compiler.filename),
        # When the serialization path is the same as the workbook
        ('{}.yml'.format(excel_compiler.filename), excel_compiler.filename),
    ):
        excel_compiler._to_text(serialization_filename)
        deserialized = excel_compiler._from_text(serialization_filename)
        assert expected == deserialized.filename

    # When the serialized data does not contain a filename we should use the
    # passed in filename to _from_text when de-serializing - for compatibility
    with open(serialization_override_path, 'r+') as f:
        # Modify our previously serialized compiler and remove the filename key
        f.data = YAML().load(f)
        f.seek(0)
        f.data.pop('filename')
        YAML().dump(f.data, f)
        f.truncate()

    deserialized = excel_compiler._from_text(serialization_override_path)
    expected = serialization_override_path.rsplit('.')[0]
    assert expected == deserialized.filename


def test_filename_extension_errors(excel_compiler, fixture_xls_path):
    with pytest.raises(ValueError, match='Unrecognized file type'):
        ExcelCompiler.from_file(excel_compiler.filename + '.xyzzy')

    with pytest.raises(ValueError, match='Only allowed one '):
        excel_compiler.to_file(file_types=('pkl', 'pickle'))

    with pytest.raises(ValueError, match='Only allowed one '):
        excel_compiler.to_file(file_types=('pkl', 'yml', 'json'))

    with pytest.raises(ValueError, match='Unknown file types: pkly'):
        excel_compiler.to_file(file_types=('pkly',))


def test_hash_matches(excel_compiler):
    assert excel_compiler.hash_matches

    excel_compiler._excel_file_md5_digest = 0
    assert not excel_compiler.hash_matches


def test_pickle_file_rebuilding(excel_compiler):

    input_addrs = ['Sheet1!A11']
    output_addrs = ['Sheet1!D1']

    excel_compiler.trim_graph(input_addrs, output_addrs)
    excel_compiler.to_file()

    pickle_name = excel_compiler.filename + '.pkl'
    yaml_name = excel_compiler.filename + '.yml'

    assert os.path.exists(pickle_name)
    old_hash = excel_compiler._compute_file_md5_digest(pickle_name)

    excel_compiler.to_file()
    assert old_hash == excel_compiler._compute_file_md5_digest(pickle_name)

    os.unlink(yaml_name)
    excel_compiler.to_file()
    new_hash = excel_compiler._compute_file_md5_digest(pickle_name)
    assert old_hash != new_hash

    shutil.copyfile(pickle_name, yaml_name)
    excel_compiler.to_file()
    assert new_hash != excel_compiler._compute_file_md5_digest(pickle_name)


def test_reset(excel_compiler):
    in_address = 'Sheet1!A1'
    out_address = 'Sheet1!D1'

    assert -0.02286 == round(excel_compiler.evaluate(out_address), 5)

    in_value = excel_compiler.cell_map[in_address].value

    excel_compiler._reset(excel_compiler.cell_map[in_address])
    assert excel_compiler.cell_map[out_address].value is None

    excel_compiler._reset(excel_compiler.cell_map[in_address])
    assert excel_compiler.cell_map[out_address].value is None

    excel_compiler.cell_map[in_address].value = in_value
    assert -0.02286 == round(excel_compiler.evaluate(out_address), 5)
    assert -0.02286 == round(excel_compiler.cell_map[out_address].value, 5)


def test_recalculate(excel_compiler):
    out_address = 'Sheet1!D1'

    assert -0.02286 == round(excel_compiler.evaluate(out_address), 5)
    excel_compiler.cell_map[out_address].value = None

    excel_compiler.recalculate()
    assert -0.02286 == round(excel_compiler.cell_map[out_address].value, 5)


def test_evaluate_from_generator(excel_compiler):
    result = excel_compiler.evaluate(
        a for a in ('trim-range!B1', 'trim-range!B2'))
    assert (24, 136) == result


def test_evaluate_empty(excel_compiler):
    assert 0 == excel_compiler.evaluate('Empty!B1')

    excel_compiler.recalculate()
    assert 0 == excel_compiler.evaluate('Empty!B1')

    input_addrs = ['Empty!C1', 'Empty!B2']
    output_addrs = ['Empty!B1', 'Empty!B2']

    excel_compiler.trim_graph(input_addrs, output_addrs)
    excel_compiler._to_text(is_json=True)
    text_excel_compiler = ExcelCompiler._from_text(
        excel_compiler.filename, is_json=True)

    assert [0, None] == text_excel_compiler.evaluate(output_addrs)
    text_excel_compiler.set_value(input_addrs[0], 10)
    assert [10, None] == text_excel_compiler.evaluate(output_addrs)

    text_excel_compiler.set_value(input_addrs[1], 20)
    assert [10, 20] == text_excel_compiler.evaluate(output_addrs)


def test_gen_graph(excel_compiler):
    excel_compiler._gen_graph('B2')

    with pytest.raises(ValueError, match='Unknown seed'):
        excel_compiler._gen_graph(None)

    with pytest.raises(NotImplementedError, match='Linked SheetNames'):
        excel_compiler._gen_graph('=[Filename.xlsx]Sheetname!A1')


def test_value_tree_str(excel_compiler):
    out_address = 'trim-range!B2'
    excel_compiler.evaluate(out_address)

    expected = [
        'trim-range!B2 = 136',
        ' trim-range!B1 = 24',
        '  trim-range!D1:E3 = ((1, 5), (2, 6), (3, 7))',
        '   trim-range!D1 = 1',
        '   trim-range!D2 = 2',
        '   trim-range!D3 = 3',
        '   trim-range!E1 = 5',
        '   trim-range!E2 = 6',
        '   trim-range!E3 = 7',
        ' trim-range!D4:E4 = ((4, 8),)',
        '  trim-range!D4 = 4',
        '  trim-range!E4 = 8',
        ' trim-range!D5 = 100'
    ]
    assert expected == list(excel_compiler.value_tree_str(out_address))


def test_value_tree_str_circular(circular_ws):

    out_address = 'Sheet1!B8'
    circular_ws.evaluate(out_address)

    expected = [
        'Sheet1!B8 = -50',
        ' Sheet1!B1 = 0',
        '  Sheet1!B2 = 0',
        '   Sheet1!A2 = 0.2',
        '   Sheet1!B1 <- cycle',
        '   Sheet1!B3 = 0',
        '  Sheet1!B3 <- cycle',
        ' Sheet1!B6 = 50',
        '  Sheet1!A5 = 50',
        '  Sheet1!A6 = 50',
        '   Sheet1!B6 <- cycle',
        '  Sheet1!B3 <- cycle',
        '  Sheet1!B5 = 0.01',
    ]
    assert expected == list(circular_ws.value_tree_str(out_address))


def test_trim_cells(excel_compiler):
    input_addrs = ['trim-range!D5']
    output_addrs = ['trim-range!B2']

    old_value = excel_compiler.evaluate(output_addrs[0])

    excel_compiler.trim_graph(input_addrs, output_addrs)
    excel_compiler._to_text(is_json=True)

    new_value = ExcelCompiler._from_text(
        excel_compiler.filename, is_json=True).evaluate(output_addrs[0])

    assert old_value == new_value


def test_trim_cells_range(excel_compiler):
    input_addrs = [AddressRange('trim-range!D4:E4')]
    output_addrs = ['trim-range!B2']

    old_value = excel_compiler.evaluate(output_addrs[0])

    excel_compiler.trim_graph(input_addrs, output_addrs)

    excel_compiler._to_text()
    excel_compiler = ExcelCompiler._from_text(excel_compiler.filename)
    assert old_value == excel_compiler.evaluate(output_addrs[0])

    excel_compiler.set_value(input_addrs[0], [5, 6], set_as_range=True)
    assert old_value - 1 == excel_compiler.evaluate(output_addrs[0])

    excel_compiler.set_value(input_addrs[0], [4, 6])
    assert old_value - 2 == excel_compiler.evaluate(output_addrs[0])

    excel_compiler.set_value(tuple(next(input_addrs[0].rows)), [5, 6])
    assert old_value - 1 == excel_compiler.evaluate(output_addrs[0])


def test_evaluate_from_non_cells(excel_compiler):
    input_addrs = ['Sheet1!A11']
    output_addrs = ['Sheet1!A11:A13', 'Sheet1!D1', 'Sheet1!B11', ]

    old_values = excel_compiler.evaluate(output_addrs)

    excel_compiler.trim_graph(input_addrs, output_addrs)

    excel_compiler.to_file(file_types='yml')
    excel_compiler = ExcelCompiler.from_file(excel_compiler.filename)
    for expected, result in zip(
            old_values, excel_compiler.evaluate(output_addrs)):
        assert tuple(flatten(expected)) == pytest.approx(tuple(flatten(result)))

    range_cell = excel_compiler.cell_map[output_addrs[0]]
    excel_compiler._reset(range_cell)
    range_value = excel_compiler.evaluate(range_cell.address)
    assert old_values[0] == range_value


def test_validate_calcs(excel_compiler, capsys):
    input_addrs = ['trim-range!D5']
    output_addrs = ['trim-range!B2']

    excel_compiler.trim_graph(input_addrs, output_addrs)
    excel_compiler.cell_map[output_addrs[0]].value = 'JUNK'
    failed_cells = excel_compiler.validate_calcs(output_addrs)

    assert {'mismatch': {
        'trim-range!B2': ('JUNK', 136, '=B1+SUM(D4:E4)+D5')}} == failed_cells

    out, err = capsys.readouterr()
    assert '' == err
    assert 'JUNK' in out


def test_validate_calcs_all_cells(basic_ws):
    formula_cells = basic_ws.formula_cells('Sheet1')
    expected = {
        AddressCell('Sheet1!B2'),
        AddressCell('Sheet1!C2'),
        AddressCell('Sheet1!B3'),
        AddressCell('Sheet1!C3'),
        AddressCell('Sheet1!B4'),
        AddressCell('Sheet1!C4')
    }
    assert expected == set(formula_cells)
    assert {} == basic_ws.validate_calcs()


def test_validate_calcs_excel_compiler(excel_compiler):
    """Find all formula cells w/ values and verify calc"""
    errors = excel_compiler.validate_calcs()
    msg = json.dumps(errors, indent=2)
    assert msg == '{}'

    errors = excel_compiler.validate_calcs('Sheet1!B1')
    msg = json.dumps(errors, indent=2)
    assert msg == '{}'

    # Missing sheets returns empty tuple
    assert len(excel_compiler.formula_cells('JUNK-Sheet!B1')) == 0


def test_validate_calcs_empty_params():
    data = [x.strip().split() for x in """
        =INDEX($G$2:$G$4,D2) =INDEX($G$2:$I$2,,D2) =INDEX($G$2:$I$2,$A$6,D2)
        =INDEX($G$2:$G$4,D3) =INDEX($G$2:$I$2,,D3) =INDEX($G$2:$I$2,$A$6,D3)
        =INDEX($G$2:$G$4,D4) =INDEX($G$2:$I$2,,D4) =INDEX($G$2:$I$2,$A$6,D4)
        =MATCH(E2,$F$2:$F$4) 0 1 a b c
        =MATCH(E3,$F$2:$F$4) 2 2 b
        =MATCH(E4,$F$2:$F$4) 4 3 c
        =IF(0,,2) =IF(1,,2) =IF(,1,2) =IF(,0,2)
    """.split('\n')[1:-1]]

    wb = Workbook()
    ws = wb.active
    ws['A2'], ws['B2'], ws['C2'] = data[0]
    ws['A3'], ws['B3'], ws['C3'] = data[1]
    ws['A4'], ws['B4'], ws['C4'] = data[2]
    ws['D2'], ws['E2'], ws['F2'], ws['G2'], ws['H2'], ws['I2'] = data[3]
    ws['D3'], ws['E3'], ws['F3'], ws['G3'] = data[4]
    ws['D4'], ws['E4'], ws['F4'], ws['G4'] = data[5]
    ws['A5'], ws['B5'], ws['C5'], ws['D5'] = data[6]

    excel_compiler = ExcelCompiler(excel=wb)

    assert (NA_ERROR, ) * 3 == excel_compiler.evaluate('Sheet!A2:C2')
    assert ('b', ) * 3 == excel_compiler.evaluate('Sheet!A3:C3')
    assert ('c', ) * 3 == excel_compiler.evaluate('Sheet!A4:C4')
    assert (2, 0, 2, 2) == excel_compiler.evaluate('Sheet!A5:D5')


def test_evaluate_entire_row_column(excel_compiler):

    value = excel_compiler.evaluate(AddressRange('Sheet1!A:A'))
    expected = excel_compiler.evaluate(AddressRange('Sheet1!A1:A18'))
    assert value == expected
    assert len(value) == 18
    assert not list_like(value[0])

    value = excel_compiler.evaluate(AddressRange('Sheet1!1:1'))
    expected = excel_compiler.evaluate(AddressRange('Sheet1!A1:D1'))
    assert value == expected
    assert len(value) == 4
    assert not list_like(value[0])

    value = excel_compiler.evaluate(AddressRange('Sheet1!A:B'))
    expected = excel_compiler.evaluate(AddressRange('Sheet1!A1:B18'))
    assert value == expected
    assert len(value) == 18
    assert len(value[0]) == 2

    value = excel_compiler.evaluate(AddressRange('Sheet1!1:2'))
    expected = excel_compiler.evaluate(AddressRange('Sheet1!A1:D2'))
    assert value == expected
    assert len(value) == 2
    assert len(value[0]) == 4

    # now from the text based file
    excel_compiler._to_text()
    text_excel_compiler = ExcelCompiler._from_text(excel_compiler.filename)

    value = text_excel_compiler.evaluate(AddressRange('Sheet1!A:A'))
    expected = text_excel_compiler.evaluate(AddressRange('Sheet1!A1:A18'))
    assert value == expected
    assert len(value) == 18
    assert not list_like(value[0])

    value = text_excel_compiler.evaluate(AddressRange('Sheet1!1:1'))
    expected = text_excel_compiler.evaluate(AddressRange('Sheet1!A1:D1'))
    assert value == expected
    assert len(value) == 4
    assert not list_like(value[0])

    value = text_excel_compiler.evaluate(AddressRange('Sheet1!A:B'))
    expected = text_excel_compiler.evaluate(AddressRange('Sheet1!A1:B18'))
    assert len(value) == 18
    assert len(value[0]) == 2
    assert value == expected

    value = text_excel_compiler.evaluate(AddressRange('Sheet1!1:2'))
    expected = text_excel_compiler.evaluate(AddressRange('Sheet1!A1:D2'))
    assert value == expected
    assert len(value) == 2
    assert len(value[0]) == 4


def test_evaluate_conditional_formatting(cond_format_ws):
    cells_addrs = [
        AddressCell('B2'),
        AddressCell('Sheet1!B3'),
        AddressRange('Sheet1!B4:B6'),
    ]
    formats = cond_format_ws.eval_conditional_formats(cells_addrs)
    formats2 = cond_format_ws.eval_conditional_formats((a for a in cells_addrs))
    assert formats == list(formats2)
    assert len(formats) == 3
    assert len(formats[2]) == 3

    # read the spreadsheet from yaml
    cond_format_ws.to_file(file_types=('yml', ))
    cond_format_ws_yaml = ExcelCompiler.from_file(
        cond_format_ws.filename + '.yml')
    cells_addrs[0] = AddressCell('Sheet1!B2')
    formats3 = cond_format_ws_yaml.eval_conditional_formats(tuple(cells_addrs))
    assert formats2 == formats3

    # read the spreadsheet from pickle
    cond_format_ws.to_file(file_types=('pkl', ))
    cond_format_ws_pkl = ExcelCompiler.from_file(
        cond_format_ws.filename + '.pkl')
    cells_addrs[0] = AddressCell('Sheet1!B2')
    formats4 = cond_format_ws_pkl.eval_conditional_formats(tuple(cells_addrs))
    assert formats2 == formats4

    formats.append(formats[2][0][0])
    formats.append(formats[2][1][0])
    formats.append(formats[2][2][0])
    del formats[2]

    color_key = {
        ('FF006100', 'FFC6EFCE'): 'grn',
        ('FF9C5700', 'FFFFEB9C'): 'yel',
        ('FF9C0006', 'FFFFC7CE'): 'red',
        (None, 'FFFFC7CE'): 'nofont',
    }

    color_map = {}
    for idx, dxf in cond_format_ws.conditional_formats.items():
        color_map[idx] = color_key[
            dxf.font and dxf.font.color.value, dxf.fill.bgColor.value]

    expected = [
        ['red'],
        ['grn', 'yel', 'red'],
        ['yel', 'red'],
        ['nofont'],
        ['yel', 'red'],
    ]
    results = [[color_map[x] for x in y] for y in formats]
    assert results == expected


def test_trim_cells_warn_address_not_found(excel_compiler):
    input_addrs = ['trim-range!D5', 'trim-range!H1']
    output_addrs = ['trim-range!B2']

    excel_compiler.evaluate(output_addrs[0])
    excel_compiler.log.warning = mock.Mock()
    excel_compiler.trim_graph(input_addrs, output_addrs)
    assert 1 == excel_compiler.log.warning.call_count


def test_trim_cells_info_buried_input(excel_compiler):
    input_addrs = ['trim-range!B1', 'trim-range!D1']
    output_addrs = ['trim-range!B2']

    excel_compiler.evaluate(output_addrs[0])
    excel_compiler.log.info = mock.Mock()
    excel_compiler.trim_graph(input_addrs, output_addrs)
    assert 2 == excel_compiler.log.info.call_count
    assert 'not a leaf node' in excel_compiler.log.info.mock_calls[1][1][0]


def test_trim_cells_exception_input_unused(excel_compiler):
    input_addrs = ['trim-range!G1']
    output_addrs = ['trim-range!B2']
    excel_compiler.evaluate(output_addrs[0])
    excel_compiler.evaluate(input_addrs[0])

    with pytest.raises(
            ValueError,
            match=' which usually means no outputs are dependant on it'):
        excel_compiler.trim_graph(input_addrs, output_addrs)


def test_compile_error_message_line_number(excel_compiler):
    input_addrs = ['trim-range!D5']
    output_addrs = ['trim-range!B2']

    excel_compiler.trim_graph(input_addrs, output_addrs)

    filename = excel_compiler.filename + '.pickle'
    excel_compiler.to_file(filename)

    excel_compiler = ExcelCompiler.from_file(filename)
    formula = excel_compiler.cell_map[output_addrs[0]].formula
    formula._python_code = '(x)'
    formula.lineno = 3000
    formula.filename = 'a_file'
    with pytest.raises(UnknownFunction, match='File "a_file", line 3000'):
        excel_compiler.evaluate(output_addrs[0])


def test_init_cell_address_error(excel):
    with pytest.raises(ValueError):
        _CellRange(ExcelWrapper.RangeData(
            AddressCell('A1'), '', ((0, ),)))


def test_cell_range_repr(excel):
    cell_range = _CellRange(ExcelWrapper.RangeData(
        AddressRange('sheet!A1:B1'), '', ((0, 0),)))
    assert 'sheet!A1:B1' == repr(cell_range)


def test_cell_repr(excel):
    cell_range = _Cell('sheet!A1', value=0)
    assert 'sheet!A1 -> 0' == repr(cell_range)


def test_gen_gexf(excel_compiler, tmpdir):
    filename = os.path.join(str(tmpdir), 'test.gexf')
    assert not os.path.exists(filename)
    excel_compiler.export_to_gexf(filename)

    # ::TODO: it would good to test this by comparing to an fixture/artifact
    assert os.path.exists(filename)


def test_gen_dot(excel_compiler, tmpdir):
    with pytest.raises(ImportError, match="Package 'pydot' is not installed"):
        excel_compiler.export_to_dot()

    import sys
    mock_imports = (
        'pydot',
    )
    for mock_import in mock_imports:
        sys.modules[mock_import] = mock.MagicMock()

    with mock.patch('networkx.drawing.nx_pydot.write_dot'):
        excel_compiler.export_to_dot()


def test_plot_graph(excel_compiler, tmpdir):
    with pytest.raises(ImportError,
                       match="Package 'matplotlib' is not installed"):
        excel_compiler.plot_graph()

    import sys
    mock_imports = (
        'matplotlib',
        'matplotlib.pyplot',
        'matplotlib.cbook',
        'matplotlib.colors',
        'matplotlib.collections',
        'matplotlib.patches',
    )
    for mock_import in mock_imports:
        sys.modules[mock_import] = mock.MagicMock()
    out_address = 'trim-range!B2'
    excel_compiler.evaluate(out_address)

    with mock.patch('pycel.excelcompiler.nx'):
        excel_compiler.plot_graph()


def test_structured_ref(excel_compiler):
    input_addrs = ['sref!F3']
    output_addrs = ['sref!B3']

    assert 15 == excel_compiler.evaluate(output_addrs[0])
    excel_compiler.trim_graph(input_addrs, output_addrs)

    assert 15 == excel_compiler.evaluate(output_addrs[0])

    excel_compiler.set_value(input_addrs[0], 11)
    assert 20 == excel_compiler.evaluate(output_addrs[0])


def test_multi_area_range_defined_name():

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 1
    ws['A2'] = 2
    ws['A3'] = 3
    ws['A4'] = 4
    ws['B1'] = '=SUM(A1,A2)'
    ws['B2'] = '=SUM(_a2,A3)'
    ws['B3'] = '=SUM(_a1_a3)'

    wb.defined_names.append(
        DefinedName(name='_a2', attr_text='Sheet!$A$4,Sheet!$A$1:$A$2'))
    wb.defined_names.append(
        DefinedName(name='_a1_a3', attr_text='Sheet!$A$1,Sheet!$A$3'))
    excel_compiler = ExcelCompiler(excel=wb)

    output_addrs = ['Sheet!B1:B3']
    assert (3, 10, 4) == excel_compiler.evaluate(output_addrs[0])
    excel_compiler.recalculate()
    assert (3, 10, 4) == excel_compiler.evaluate(output_addrs[0])


def test_unbounded_countifs():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 1
    ws['A2'] = 2
    ws['A3'] = 3
    ws['A4'] = 4
    ws['A5'] = 5
    ws['B1'] = 1
    ws['B2'] = 2
    ws['B3'] = 3
    ws['B4'] = 4
    ws['B5'] = 5
    ws['C1'] = '=COUNTIFS(B:B,">3")'
    ws['C2'] = '=SUMIFS(A:A,B:B,">3")'
    excel_compiler = ExcelCompiler(filename='test_unbounded_countifs', excel=wb)

    output_addrs = 'Sheet!C1', 'Sheet!C2'
    assert (2, 9) == excel_compiler.evaluate(output_addrs)
    excel_compiler.recalculate()
    assert (2, 9) == excel_compiler.evaluate(output_addrs)

    # read the spreadsheet from pickle
    excel_compiler.to_file(file_types=('pickle', ))
    excel_compiler = ExcelCompiler.from_file(excel_compiler.filename)

    # test evaluation
    assert (2, 9) == excel_compiler.evaluate(output_addrs)
    excel_compiler.recalculate()
    assert (2, 9) == excel_compiler.evaluate(output_addrs)


def test_validate_count():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 0
    ws['A2'] = 1
    ws['A3'] = 1.1
    ws['A4'] = '1.1'
    ws['A5'] = True
    ws['A6'] = False
    ws['A7'] = 'A'
    ws['A8'] = 'TRUE'
    ws['A9'] = 'FALSE'
    ws['B1'] = '=COUNT(A2)'
    ws['B2'] = '=COUNT(A2:A3)'
    ws['B3'] = '=COUNT(A2:A3,A3)'
    ws['B4'] = '=COUNT(A1:A9,A4,A5,A6,A7,A8,A9)'

    excel_compiler = ExcelCompiler(excel=wb)
    assert excel_compiler.evaluate('Sheet!B1:B4') == (1, 2, 3, 3)

    # Test missing calcPr in WorkbookPackage
    wb.calculation = None
    excel_compiler = ExcelCompiler(excel=wb)
    assert excel_compiler.evaluate('Sheet!B1:B4') == (1, 2, 3, 3)


@pytest.mark.parametrize(
    'msg, formula', (
        ("Function XYZZY is not implemented. "
         "XYZZY is not a known Excel function", '=xyzzy()'),
        ("Function PLUGH is not implemented. "
         "PLUGH is not a known Excel function\n"
         "Function XYZZY is not implemented. "
         "XYZZY is not a known Excel function", '=xyzzy() + plugh()'),
        ('Function ARABIC is not implemented. '
         'ARABIC is in the "Math and trigonometry" group, '
         'and was introduced in Excel 2013',
         '=ARABIC()'),
    )
)
def test_unknown_functions(fixture_dir, msg, formula):
    excel_compiler = ExcelCompiler.from_file(
        os.path.join(fixture_dir, 'fixture.xlsx.yml'))

    address = AddressCell('s!A1')
    excel_compiler.cell_map[str(address)] = _Cell(
        address, None, formula, excel_compiler.excel
    )
    with pytest.raises(UnknownFunction, match=msg):
        excel_compiler.evaluate(address)

    result = excel_compiler.validate_calcs([address])
    assert 'not-implemented' in result
    assert len(result['not-implemented']) == 1


def test_evaluate_exceptions(fixture_dir):
    excel_compiler = ExcelCompiler.from_file(
        os.path.join(fixture_dir, 'fixture.xlsx.yml'))

    address = AddressCell('s!A1')
    excel_compiler.cell_map[str(address)] = _Cell(
        address, None, '=__REF__("s!A2")', excel_compiler.excel
    )
    address = AddressCell('s!A2')
    excel_compiler.cell_map[str(address)] = _Cell(
        address, None, '=$', excel_compiler.excel
    )

    with pytest.raises(FormulaParserError):
        excel_compiler.evaluate(address)

    result = excel_compiler.validate_calcs(address)
    assert 'exceptions' in result
    assert len(result['exceptions']) == 1


def test_evaluate_empty_intersection(fixture_dir):
    excel_compiler = ExcelCompiler.from_file(
        os.path.join(fixture_dir, 'fixture.xlsx.yml'))

    address = AddressCell('s!A1')
    excel_compiler.cell_map[str(address)] = _Cell(
        address, None, '=_R_(str(_REF_("s!A1:A2") & _REF_("s!B1:B2")))',
        excel_compiler.excel
    )
    assert excel_compiler.evaluate(address) == NULL_ERROR


def test_plugins(excel_compiler):

    input_addrs = ['Sheet1!A11']
    output_addrs = ['Sheet1!D1']
    excel_compiler.trim_graph(input_addrs, output_addrs)
    d1 = -0.022863768173008364

    excel_compiler.recalculate()
    assert pytest.approx(d1) == excel_compiler.evaluate('Sheet1!D1')

    def calc_and_check():
        excel_compiler._eval = None
        excel_compiler.cell_map['Sheet1!D1'].formula.compiled_lambda = None
        excel_compiler.recalculate()
        assert pytest.approx(d1) == excel_compiler.evaluate('Sheet1!D1')

    with mock.patch('pycel.excelformula.ExcelFormula.default_modules', ()):
        with pytest.raises(UnknownFunction):
            calc_and_check()

    with mock.patch('pycel.excelformula.ExcelFormula.default_modules', ()):
        excel_compiler._plugin_modules = ('pycel.excellib', )
        calc_and_check()

    with mock.patch('pycel.excelformula.ExcelFormula.default_modules', ()):
        excel_compiler._plugin_modules = 'pycel.excellib'
        calc_and_check()

    with mock.patch('pycel.excelformula.ExcelFormula.default_modules',
                    ('pycel.excellib', )):
        excel_compiler._plugin_modules = None
        calc_and_check()

    with mock.patch('pycel.excelformula.ExcelFormula.default_modules', ()):
        with pytest.raises(UnknownFunction):
            calc_and_check()


@pytest.mark.parametrize(
    'a2, b3, iters, result', (
        (0.2, 100, 3, 16.8),
        (0.2, 100, 4, 16.64),
        (0.2, 100, 5, 16.672),
        (0.2, 200, 3, 33.6),
        (0.2, 200, 4, 33.28),
        (0.2, 200, 5, 33.344),
        (0.2, 500, 3, 84),
        (0.2, 500, 4, 83.2),
        (0.2, 500, 5, 83.36),
        (0.1234, 500, 3, 55.025760452),
    )
)
def test_validate_circular_referenced_iters(
        circular_ws, a2, b3, iters, result):
    circular_ws.evaluate(['Sheet1!B2', 'Sheet1!B6'], iterations=1)
    circular_ws.set_value('Sheet1!A2', a2)
    circular_ws.set_value('Sheet1!B2', 0)
    circular_ws.set_value('Sheet1!B3', b3)
    val = circular_ws.evaluate('Sheet1!B2', iterations=iters, tolerance=1e-100)
    assert val == pytest.approx(result)


@pytest.mark.parametrize(
    'start, inc, tol, result', (
        (50, 0.1, 0.09, 40),
        (50, 0.1, 0.1, 49.9),
        (50, 0.1, 0.11, 49.9),
    )
)
def test_validate_circular_referenced_tol(
        circular_ws, start, inc, tol, result):
    circular_ws.evaluate(['Sheet1!B2', 'Sheet1!B6'], iterations=1)
    circular_ws.set_value('Sheet1!B3', 100)
    circular_ws.set_value('Sheet1!A5', start)
    circular_ws.set_value('Sheet1!B5', inc)
    circular_ws.set_value('Sheet1!A6', start)
    circular_ws.set_value('Sheet1!B6', start)
    assert circular_ws.evaluate('Sheet1!B6', iterations=100, tolerance=tol
                                ) == pytest.approx(result)


def test_validate_circular_referenced(circular_ws):
    b6_expect = pytest.approx(49.92)
    b8_expect = pytest.approx(33.41312)
    circular_ws.evaluate('Sheet1!B8', iterations=1)

    circular_ws.set_value('Sheet1!B3', 0)
    b8 = circular_ws.evaluate('Sheet1!B8', iterations=5000, tolerance=1e-20)
    assert b8 == -50
    circular_ws.set_value('Sheet1!B3', 100)
    b8 = circular_ws.evaluate('Sheet1!B8', iterations=5000, tolerance=0.01)
    assert b8 == b8_expect

    circular_ws.set_value('Sheet1!B3', 0)
    b6, b8 = circular_ws.evaluate(
        ['Sheet1!B6', 'Sheet1!B8'], iterations=5000, tolerance=1e-20)
    assert (b6, b8) == (50, -50)
    circular_ws.set_value('Sheet1!B3', 100)
    b6, b8 = circular_ws.evaluate(
        ['Sheet1!B6', 'Sheet1!B8'], iterations=5000, tolerance=0.01)
    assert b6 == b6_expect
    assert b8 == b8_expect

    circular_ws.set_value('Sheet1!B3', 0)
    b6, b8 = circular_ws.evaluate(
        ['Sheet1!B6', 'Sheet1!B8'], iterations=5000, tolerance=1e-20)
    assert (b6, b8) == (50, -50)
    circular_ws.set_value('Sheet1!B3', 100)
    b8, b6 = circular_ws.evaluate(
        ['Sheet1!B8', 'Sheet1!B6'], iterations=5000, tolerance=0.01)
    assert b6 == b6_expect
    assert b8 == b8_expect

    # round trip cycle params through text file
    circular_ws.to_file(file_types='yml')
    excel_compiler = ExcelCompiler.from_file(circular_ws.filename)
    excel_compiler.set_value('Sheet1!B3', 0)
    b6, b8 = excel_compiler.evaluate(
        ['Sheet1!B6', 'Sheet1!B8'], iterations=5000, tolerance=1e-20)
    assert (b6, b8) == (50, -50)
    excel_compiler.set_value('Sheet1!B3', 100)
    b8, b6 = excel_compiler.evaluate(
        ['Sheet1!B8', 'Sheet1!B6'], iterations=5000, tolerance=0.01)
    assert b6 == b6_expect
    assert b8 == b8_expect


def test_circular_mismatch_warning(
        fixture_xls_path, fixture_xls_path_circular):

    with mock.patch('pycel.excelcompiler.pycel_logger') as log:
        assert log.warning.call_count == 0

        ExcelCompiler(fixture_xls_path, cycles=False)
        assert log.warning.call_count == 0

        ExcelCompiler(fixture_xls_path, cycles=True)
        assert log.warning.call_count == 1

        ExcelCompiler(fixture_xls_path_circular, cycles=False)
        assert log.warning.call_count == 2

        ExcelCompiler(fixture_xls_path_circular, cycles=True)
        assert log.warning.call_count == 2
