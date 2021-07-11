# -*- coding: UTF-8 -*-
#
# Copyright 2011-2019 by Dirk Gorissen, Stephen Rauch and Contributors
# All rights reserved.
# This file is part of the Pycel Library, Licensed under GPLv3 (the 'License')
# You may not use this work except in compliance with the License.
# You may obtain a copy of the Licence at:
#   https://www.gnu.org/licenses/gpl-3.0.en.html

import collections
import itertools as it
import operator
import re
import threading

import numpy as np
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import (
    get_column_letter,
    quote_sheetname,
    range_boundaries as openpyxl_range_boundaries,
)


ERROR_CODES = frozenset(Tokenizer.ERROR_CODES)
DIV0 = '#DIV/0!'
EMPTY = '#EMPTY!'
VALUE_ERROR = '#VALUE!'
NUM_ERROR = '#NUM!'
NA_ERROR = '#N/A'
NAME_ERROR = "#NAME?"
NULL_ERROR = "#NULL!"
REF_ERROR = "#REF!"

R1C1_ROW_RE_STR = r"R(\[-?\d+\]|\d+)?"
R1C1_COL_RE_STR = r"C(\[-?\d+\]|\d+)?"
R1C1_COORD_RE_STR = f"(?P<row>{R1C1_ROW_RE_STR})?(?P<col>{R1C1_COL_RE_STR})?"
R1C1_COORDINATE_RE = re.compile('^' + R1C1_COORD_RE_STR + '$', re.VERBOSE)

R1C1_RANGE_EXPR = f"""
(?P<min_row>{R1C1_ROW_RE_STR})?
(?P<min_col>{R1C1_COL_RE_STR})?
(:(?P<max_row>{R1C1_ROW_RE_STR})?
(?P<max_col>{R1C1_COL_RE_STR})?)?
"""

R1C1_RANGE_RE = re.compile('^' + R1C1_RANGE_EXPR + '$', re.VERBOSE)

TABLE_REF_RE = re.compile(r"^(?P<table_name>[^[]+)\[(?P<table_selector>.*)\]$")

TABLE_SELECTOR_RE = re.compile(
    r"^(?P<row_or_column>[^[]+)$|"
    r"^@\[(?P<this_row_column>[^[]*)\]$|"
    r"^ *(?P<rows>(\[([^\]]+)\] *, *)*)"
    r"(\[(?P<start_col>[^\]]+)\] *: *)?"
    r"(\[(?P<end_col>.+)\] *)?$")

QUESTION_MARK_RE = re.compile(r'\?(?<!~)')
STAR_RE = re.compile(r'\*(?<!~)')

MAX_COL = 16384
MAX_ROW = 1048576

VALID_R1C1_RANGE_ITEM_COMBOS = {
    (0, 1, 0, 1),
    (1, 0, 1, 0),
    (1, 1, 1, 1),
}

OPERATORS = {
    '': operator.eq,
    '=': operator.eq,
    '<': operator.lt,
    '>': operator.gt,
    '<=': operator.le,
    '>=': operator.ge,
    '<>': operator.ne,
}

OPERATORS_RE = re.compile('^(?P<oper>(=|<>|<=?|>=?))?(?P<value>.*)$')

PYTHON_AST_OPERATORS = {
    'Eq': operator.eq,
    'Lt': operator.lt,
    'Gt': operator.gt,
    'LtE': operator.le,
    'GtE': operator.ge,
    'NotEq': operator.ne,
    'Add': operator.add,
    'Sub': operator.sub,
    'UAdd': operator.pos,
    'USub': operator.neg,
    'Mult': operator.mul,
    'Div': operator.truediv,
    'FloorDiv': operator.floordiv,
    'Mod': operator.mod,
    'Pow': operator.pow,
    'LShift': operator.lshift,
    'RShift': operator.rshift,
    'BitOr': operator.or_,
    'BitXor': operator.xor,
    'BitAnd': operator.and_,
    'MatMult': operator.matmul,
}

COMPARISION_OPS = frozenset(('Eq', 'Lt', 'Gt', 'LtE', 'GtE', 'NotEq'))


AddressSize = collections.namedtuple('AddressSize', 'height width')


class PyCelException(Exception):
    """Base class for PyCel errors"""


class AddressMixin:

    def __str__(self):
        return self.address

    @property
    def has_sheet(self):
        """Does the address have a sheet?"""
        return bool(self.sheet)

    @staticmethod
    def quote_sheet(sheet):
        if ' ' in sheet:
            sheet = quote_sheetname(sheet)
        return sheet

    @property
    def quoted_address(self):
        """requote the sheetname if going to include in formulas"""
        return f"{self.quote_sheet(self.sheet)}!{self.coordinate}"

    @property
    def abs_address(self):
        return f"{self.quote_sheet(self.sheet)}!{self.abs_coordinate}"

    @property
    def sort_key(self):
        return self.sheet, self.col_idx, self.row

    def _union_instersection(self, other, min_, max_):
        """Assumes rectangular only"""
        if not is_address(other):
            other = AddressRange.create(other)
        if self.sheet and other.sheet and self.sheet != other.sheet:
            return VALUE_ERROR

        min_col_idx = min_(self.col_idx, other.col_idx)
        min_row = min_(self.row, other.row)

        max_col_idx = max_(self.col_idx + self.size.width,
                           other.col_idx + other.size.width) - 1
        max_row = max_(self.row + self.size.height,
                       other.row + other.size.height) - 1

        if max_col_idx < min_col_idx or max_row < min_row:
            return NULL_ERROR

        elif max_col_idx == min_col_idx and max_row == min_row:
            return AddressCell((min_col_idx, min_row, max_col_idx, max_row),
                               sheet=self.sheet or other.sheet)
        else:
            return AddressRange((min_col_idx, min_row, max_col_idx, max_row),
                                sheet=self.sheet or other.sheet)

    def __pow__(self, other):
        return self._union_instersection(other, min, max)

    def __rpow__(self, other):
        return self._union_instersection(other, min, max)

    def __and__(self, other):
        return self._union_instersection(other, max, min)

    def __rand__(self, other):
        return self._union_instersection(other, max, min)


class AddressRange(collections.namedtuple(
        'Address', 'address sheet start end coordinate'), AddressMixin):
    """ Helper class for constructing, validating and accessing Range Addresses

    **Tuple Attributes:**

    .. py:attribute:: address

        `AddressRange` as a string

    .. py:attribute:: sheet

        Sheet name

    .. py:attribute:: start

        `AddressCell` for upper left corner of `AddressRange`

    .. py:attribute:: end

        `AddressCell` for lower right corner of `AddressRange`

    .. py:attribute:: coordinate

        Address without the sheetname

    **Non-tuple Attributes:**

    """

    def __new__(cls, address, *args, sheet=''):
        if args:
            return super(AddressRange, cls).__new__(cls, address, *args)

        if isinstance(address, str):
            return cls.create(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            return AddressCell(address, sheet=sheet)

        elif isinstance(address, AddressRange):
            if not sheet or sheet == address.sheet:
                return address

            elif not address.sheet:
                start = AddressCell(address.start.coordinate, sheet=sheet)
                end = AddressCell(address.end.coordinate, sheet=sheet)

            else:
                raise ValueError(f"Mismatched sheets '{address}' and '{sheet}'")

        else:
            assert (isinstance(address, tuple) and 4 == len(address) and
                    None in address or address[0:2] != address[2:]), \
                f"AddressRange expected a range '{address}'"

            start_col, start_row, end_col, end_row = address
            start = AddressCell((start_col, start_row, start_col, start_row), sheet=sheet)
            end = AddressCell((end_col, end_row, end_col, end_row), sheet=sheet)

        coordinate = f'{start.coordinate}:{end.coordinate}'

        format_str = '{0}!{1}' if sheet else '{1}'
        return super(AddressRange, cls).__new__(
            cls, format_str.format(sheet, coordinate),
            sheet, start, end, coordinate)

    def __contains__(self, address):
        address = AddressCell(address)
        return (self.start.row <= address.row <= self.end.row and
                self.start.col_idx <= address.col_idx <= self.end.col_idx)

    @property
    def col_idx(self):
        """col_idx for left column"""
        return self.start.col_idx

    @property
    def row(self):
        """top row"""
        return self.start.row

    @property
    def abs_coordinate(self):
        return f'{self.start.abs_coordinate}:{self.end.abs_coordinate}'

    # Is this address a range?
    is_range = True

    @property
    def is_unbounded_range(self):
        """Is this address an unbounded range?"""
        rows, cols = self.size
        return rows == MAX_ROW or cols == MAX_COL

    @property
    def size(self):
        """Range dimensions"""
        if not hasattr(self, '_size'):
            if 0 in (self.end.row, self.start.row):
                height = MAX_ROW
            else:
                height = self.end.row - self.start.row + 1

            if 0 in (self.end.col_idx, self.start.col_idx):
                width = MAX_COL
            else:
                width = self.end.col_idx - self.start.col_idx + 1

            self._size = AddressSize(height, width)
        return self._size

    @property
    def rows(self):
        """Get each address for every cell, yields one row at a time."""
        col_range = self.start.col_idx, self.end.col_idx + 1
        for row in range(self.start.row, self.end.row + 1):
            yield (AddressCell((col, row, col, row), sheet=self.sheet)
                   for col in range(*col_range))

    @property
    def cols(self):
        """Get each address for every cell, yields one column at a time."""
        col_range = self.start.col_idx, self.end.col_idx + 1
        for col in range(*col_range):
            yield (AddressCell((col, row, col, row), sheet=self.sheet)
                   for row in range(self.start.row, self.end.row + 1))

    def address_at_offset(self, row_inc=0, col_inc=0):
        return self.start.address_at_offset(row_inc=row_inc, col_inc=col_inc)

    @property
    def resolve_range(self):
        """Return nested tuples with an AddressCell for each element"""
        assert not self.is_unbounded_range
        return tuple(tuple(row) for row in self.rows)

    @classmethod
    def create(cls, address, sheet='', cell=None):
        """ Factory method.

        Able to construct R1C1, defined names, and structured references
        style addresses, if passed a `excelcompiler._Cell`.

        :param address: str, AddressRange, AddressCell
        :param sheet: sheet for address, if not included
        :param cell: `excelcompiler._Cell` reference
        :return: `AddressRange or AddressCell`
        """

        if isinstance(address, AddressRange):
            return AddressRange(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            return AddressCell(address, sheet=sheet)

        elif address in ERROR_CODES:
            return address

        sheetname, addr = split_sheetname(address, sheet=sheet)
        addr_tuple, sheetname = range_boundaries(
            addr, sheet=sheetname, cell=cell)

        if isinstance(addr_tuple, AddressMultiAreaRange):
            return addr_tuple
        elif None in addr_tuple or addr_tuple[0:2] != addr_tuple[2:]:
            return AddressRange(addr_tuple, sheet=sheetname)
        else:
            return AddressCell(addr_tuple, sheet=sheetname)


class AddressCell(collections.namedtuple(
        'AddressCell', 'address sheet col_idx row coordinate'), AddressMixin):
    """ Helper class for constructing, validating and accessing Cell Addresses

    **Tuple Attributes:**

    .. py:attribute:: address

        `AddressRange` as a string

    .. py:attribute:: sheet

        Sheet name

    .. py:attribute:: col_idx

        Column number as a 1 based index

    .. py:attribute:: row

        Row number as a 1 based index

    .. py:attribute:: coordinate

        Address without the sheetname

    **Non-tuple Attributes:**

    """

    def __new__(cls, address, *args, sheet=''):
        if args:
            return super(AddressCell, cls).__new__(cls, address, *args)

        if isinstance(address, str):
            return cls.create(address, sheet=sheet)

        elif isinstance(address, AddressCell):
            if not sheet or sheet == address.sheet:
                return address

            elif not address.sheet:
                col_idx, row, coordinate = address[2:5]

            else:
                raise ValueError(f"Mismatched sheets '{address}' and '{sheet}'")

        else:
            assert (isinstance(address, tuple) and 4 == len(address) and
                    None not in address or address[0:2] == address[2:]), \
                f"AddressCell expected a cell '{address}'"

            col_idx, row = (a or 0 for a in address[:2])
            column = (col_idx or '') and get_column_letter(col_idx)
            coordinate = f'{column}{row or ""}'

        if sheet:
            format_str = '{0}!{1}'
        else:
            format_str = '{1}'

        return super(AddressCell, cls).__new__(
            cls, format_str.format(sheet, coordinate),
            sheet, col_idx, row, coordinate)

    def __contains__(self, address):
        return self == AddressCell(address)

    # Is this address a range?
    is_range = False

    # Is this address an unbounded range?"""
    is_unbounded_range = False

    size = AddressSize(1, 1)

    @property
    def column(self):
        """column letter"""
        return (self.col_idx or '') and get_column_letter(self.col_idx)

    def inc_col(self, inc):
        """ Generate an address offset by `inc` columns.

        :param inc: integer number of columns to offset by
        """
        return (self.col_idx + inc - 1) % MAX_COL + 1

    def inc_row(self, inc):
        """ Generate an address offset by `inc` rows.

        :param inc: integer number of rows to offset by
        """
        return (self.row + inc - 1) % MAX_ROW + 1

    @property
    def abs_coordinate(self):
        return f'${self.column}${self.row}'

    def address_at_offset(self, row_inc=0, col_inc=0):
        """ Construct an `AddressCell` offset from the address

        :param row_inc: Number of rows to offset.
        :param col_inc: Number of columns to offset
        :return: `AddressCell`
        """
        new_col = self.inc_col(col_inc)
        new_row = self.inc_row(row_inc)
        return AddressCell((new_col, new_row, new_col, new_row),
                           sheet=self.sheet)

    @property
    def start(self):
        return self

    @property
    def end(self):
        return self

    @property
    def resolve_range(self):
        """Return nested tuples with an AddressCell for each element"""
        return (self, ),

    @classmethod
    def create(cls, address, sheet='', cell=None):
        """ Factory method.

        Able to construct R1C1, defined names, and structured references
        style addresses, if passed a `excelcomppiler._Cell`.

        :param address: str, AddressRange, AddressCell
        :param sheet: sheet for address, if not included
        :param cell: `excelcompiler._Cell` reference
        :return: `AddressCell`
        """
        addr = AddressRange.create(address, sheet=sheet, cell=cell)
        if not isinstance(addr, AddressCell):
            raise ValueError(f"{address} is not a valid coordinate")
        return addr


class AddressMultiAreaRange(tuple):
    """Multi-Area Address Range"""

    def __str__(self):
        return ','.join(str(addr) for addr in self)

    def __contains__(self, address):
        address = AddressCell(address)
        return any(address in addr for addr in self)

    # Is this address a range?
    is_range = True

    @property
    def is_unbounded_range(self):
        """Is this address an unbounded range?"""
        return any(addr.is_unbounded_range for addr in self
                   if isinstance(addr, AddressRange))

    @property
    def resolve_range(self):
        """Return nested tuples with an AddressCell for each element"""
        return it.chain.from_iterable(addr.resolve_range for addr in self)


def is_address(addr):
    return isinstance(addr, (AddressCell, AddressRange))


def is_array_arg(arg):
    return isinstance(arg, tuple) and not is_address(arg) and isinstance(arg[0], tuple)


def has_array_arg(*args):
    return any(is_array_arg(a) for a in args)


def unquote_sheetname(sheetname):
    """
    Remove quotes from around, an embedded "''" in, quoted sheetnames

    sheetnames with special characters are quoted in formulas
    This is the inverse of openpyxl.utils.quote_sheetname
    """
    if sheetname.startswith("'") and sheetname.endswith("'"):
        sheetname = sheetname[1:-1].replace("''", "'")
    return sheetname


def split_sheetname(address, sheet=''):
    sh = ''
    if '!' in address:
        sh, address_part = address.split('!', maxsplit=1)

        # Remove redundant sheet references and deal with inner quotes
        redundant_sheet = unquote_sheetname(sh).replace("'", "''")
        address_part = address_part.replace(f"'{redundant_sheet}'!", '')

        if '!' in address_part:
            raise NotImplementedError(f"Non-rectangular formulas: {address}")
        sh = unquote_sheetname(sh)
        address = address_part

        if sh and sheet and sh != sheet:
            raise ValueError(f"Mismatched sheets '{sh}' and '{sheet}'")

    return sheet or sh, address


def structured_reference_boundaries(address, cell=None):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   Using-structured-references-with-Excel-tables-
    #   F5ED2452-2337-4F71-BED3-C8AE6D2B276E

    match = TABLE_REF_RE.match(address)
    if not match:
        return None

    if cell is None:
        raise PyCelException(f"Must pass cell for Structured Reference {address}")

    name = match.group('table_name')
    table, sheet = cell.excel.table(name)

    if table is None:
        raise PyCelException(f"Table {name} not found for Structured Reference: {address}")

    boundaries = openpyxl_range_boundaries(table.ref)
    assert None not in boundaries

    selector = match.group('table_selector')

    if not selector:
        # all columns and the data rows
        rows, start_col, end_col = None, None, None

    else:
        selector_match = TABLE_SELECTOR_RE.match(selector)
        if selector_match is None:
            raise PyCelException(f"Unknown Structured Reference Selector: {selector}")

        row_or_column = selector_match.group('row_or_column')
        this_row_column = selector_match.group('this_row_column')

        if row_or_column:
            rows = start_col = None
            end_col = row_or_column

        elif this_row_column:
            rows = '#This Row'
            start_col = None
            end_col = this_row_column

        else:
            rows = selector_match.group('rows')
            start_col = selector_match.group('start_col')
            end_col = selector_match.group('end_col')

            if not rows:
                rows = None

            else:
                assert '[' in rows
                rows = [r.split(']')[0] for r in rows.split('[')[1:]]
                if len(rows) != 1:
                    # not currently supporting multiple row selects
                    raise PyCelException(f"Unknown Structured Reference Rows: {address}")

                rows = rows[0]

        if end_col.startswith('#'):
            # end_col collects the single field case
            assert rows is None and start_col is None
            rows = end_col
            end_col = None

        elif end_col.startswith('@'):
            rows = '#This Row'
            end_col = end_col[1:]
            if len(end_col) == 0:
                end_col = start_col

    if rows is None:
        # skip the headers and footers
        min_row = boundaries[1] + (
            table.headerRowCount if table.headerRowCount else 0)
        max_row = boundaries[3] - (
            table.totalsRowCount if table.totalsRowCount else 0)

    else:
        if rows == '#All':
            min_row, max_row = boundaries[1], boundaries[3]

        elif rows == '#Data':
            min_row = boundaries[1] + (
                table.headerRowCount if table.headerRowCount else 0)
            max_row = boundaries[3] - (
                table.totalsRowCount if table.totalsRowCount else 0)

        elif rows == '#Headers':
            min_row = boundaries[1]
            max_row = boundaries[1] + (
                table.headerRowCount if table.headerRowCount else 0) - 1

        elif rows == '#Totals':
            min_row = boundaries[3] - (
                table.totalsRowCount if table.totalsRowCount else 0) + 1
            max_row = boundaries[3]

        elif rows == '#This Row':
            # ::TODO:: If not in a data row, return #VALUE! How to do this?
            min_row = max_row = cell.address.row

        else:
            raise PyCelException(f"Unknown Structured Reference Rows: {rows}")

    if end_col is None:
        # all columns
        min_col_idx, max_col_idx = boundaries[0], boundaries[2]

    else:
        # a specific column
        column_idx = next((idx for idx, c in enumerate(table.tableColumns)
                           if c.name == end_col), None)
        if column_idx is None:
            raise PyCelException(
                f"Column {end_col} not found for Structured Reference: {address}")
        max_col_idx = boundaries[0] + column_idx

        if start_col is None:
            min_col_idx = max_col_idx

        else:
            column_idx = next((idx for idx, c in enumerate(table.tableColumns)
                               if c.name == start_col), None)
            if column_idx is None:
                raise PyCelException(
                    f"Column {start_col} not found for Structured Reference: {address}")
            min_col_idx = boundaries[0] + column_idx

    if min_row > max_row or min_col_idx > max_col_idx:
        raise PyCelException(f"Columns out of order : {address}")

    return (min_col_idx, min_row, max_col_idx, max_row), sheet


def range_boundaries(address, cell=None, sheet=None):
    try:
        # if this is normal reference then just use the openpyxl converter
        boundaries = openpyxl_range_boundaries(address)
        if None not in boundaries or ':' in address:
            return boundaries, sheet
    except ValueError:
        pass

    # test for R1C1 style address
    boundaries = r1c1_boundaries(address, cell=cell, sheet=sheet)
    if boundaries:
        return boundaries

    # Try to see if the is a structured table reference
    boundaries = structured_reference_boundaries(address, cell=cell)
    if boundaries:
        return boundaries

    # Try to see if this is a defined name
    name_addr = cell and cell.excel and cell.excel.defined_names.get(address)
    if name_addr:
        if len(name_addr) == 1:
            return openpyxl_range_boundaries(name_addr[0][0]), name_addr[0][1]
        else:
            return AddressMultiAreaRange(tuple(
                AddressRange(range_alias, sheet=worksheet)
                for range_alias, worksheet in name_addr)), None

    addrs = address.split(':')
    if len(addrs) > 2:
        # Multi colon range resolves to rectangle containing all nodes
        try:
            nodes = tuple(AddressRange.create(addr, cell=cell, sheet=sheet)
                          for addr in addrs)

            min_col_idx = min(n.col_idx for n in nodes)
            max_col_idx = max((n.col_idx + n.size.width - 1) for n in nodes)
            min_row = min(n.row for n in nodes)
            max_row = max((n.row + n.size.height - 1) for n in nodes)

            sheets = {n.sheet for n in nodes if n.sheet}
            if not sheet:
                sheet = next(iter(sheets), None)
            assert not sheets or sheets == {sheet}

            return (min_col_idx, min_row, max_col_idx, max_row), sheet
        except ValueError:
            pass

    raise ValueError(f"{address} is not a valid coordinate or range")


def r1c1_boundaries(address, cell=None, sheet=None):
    """
    R1C1 reference style

    You can also use a reference style where both the rows and the columns on
    the worksheet are numbered. The R1C1 reference style is useful for
    computing row and column positions in macros. In the R1C1 style, Excel
    indicates the location of a cell with an "R" followed by a row number
    and a "C" followed by a column number.

    Reference   Meaning

    R[-2]C      A relative reference to the cell two rows up and in
                the same column

    R[2]C[2]    A relative reference to the cell two rows down and
                two columns to the right

    R2C2        An absolute reference to the cell in the second row and
                in the second column

    R[-1]       A relative reference to the entire row above the active cell

    R           An absolute reference to the current row as part of a range

    """

    # test for R1C1 style address
    m = R1C1_RANGE_RE.match(address)

    if not m:
        return None

    def from_relative_to_absolute(r1_or_c1):
        def require_cell():
            assert cell is not None, \
                f"Must pass a cell to decode a relative address {address}"

        if not r1_or_c1.endswith(']'):
            if len(r1_or_c1) > 1:
                return int(r1_or_c1[1:])

            else:
                require_cell()
                if r1_or_c1[0].upper() == 'R':
                    return cell.row
                else:
                    return cell.col_idx

        else:
            require_cell()
            if r1_or_c1[0].lower() == 'r':
                return (cell.row + int(r1_or_c1[2:-1]) - 1) % MAX_ROW + 1
            else:
                return (cell.col_idx + int(r1_or_c1[2:-1]) - 1) % MAX_COL + 1

    min_col, min_row, max_col, max_row = (
        g if g is None else from_relative_to_absolute(g) for g in (
            m.group(n) for n in ('min_col', 'min_row', 'max_col', 'max_row')
        )
    )

    items_present = (min_col is not None, min_row is not None,
                     max_col is not None, max_row is not None)

    is_range = ':' in address
    if (is_range and items_present not in VALID_R1C1_RANGE_ITEM_COMBOS or
            not is_range and sum(items_present) < 2):
        raise ValueError(f"{address} is not a valid coordinate or range")

    if min_col is not None:
        min_col = min_col

    if min_row is not None:
        min_row = min_row

    if max_col is not None:
        max_col = max_col
    else:
        max_col = min_col

    if max_row is not None:
        max_row = max_row
    else:
        max_row = min_row

    return (min_col, min_row, max_col, max_row), sheet


class _ArrayFormulaContext:
    """ When evaluating array like data, need to know the context
        that the result will end up in
    """
    _ns = threading.local()

    @property
    def ns(self):
        if not hasattr(self._ns, 'ctx_addresses'):
            self._ns.ctx_addresses = [False]
            self._ns._ctx_address = None
        return self._ns

    def __bool__(self):
        return bool(self.ctx_address)

    def __call__(self, address):
        self.ns._ctx_address = address
        return self

    def __enter__(self):
        self.ns.ctx_addresses.append(self.ns._ctx_address)
        self.ns._ctx_address = None

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.ns.ctx_addresses.pop()

    @property
    def ctx_address(self):
        return self.ns.ctx_addresses[-1]

    def fit_to_range(self, result):
        """Expand/Contract an answer to fill a range"""
        ctx_address = self.ctx_address
        if ctx_address is not None:

            if list_like(result):
                # results are either scalar or rectangular array
                assert list_like(result[0])
                result_size = AddressSize(len(result), len(result[0]))
            else:
                result_size = AddressSize(1, 1)
                result = ((result, ), )

            ctx_size = ctx_address.size

            # if result is one col wide and target is wider, then expand columns
            if result_size.width == 1 and ctx_size.width != 1:
                result = tuple(r * ctx_size.width for r in result)

            # if result is wider than target, trim it
            elif result_size.width > ctx_size.width:
                result = tuple(row[:ctx_size.width] for row in result)

            # if result is narrower than target, fill w/ NA
            elif result_size.width < ctx_size.width:
                fill = (NA_ERROR, ) * (ctx_size.width - result_size.width)
                result = tuple(row + fill for row in result)

            # if result is one row high and target is taller, then expand rows
            if result_size.height == 1 and ctx_size.height != 1:
                result *= ctx_size.height

            # if result is taller than target, trim it
            elif result_size.height > ctx_size.height:
                result = result[:ctx_size.height]

            # if result is shorter than target, fill w/ NA
            elif result_size.height < ctx_size.height:
                fill = ((NA_ERROR, ) * ctx_size.width, )
                result += fill * (ctx_size.height - result_size.height)

        return result


in_array_formula_context = _ArrayFormulaContext()


def flatten(data, coerce=lambda x: x):
    """ flatten items, converting top level items as needed

    :param data: data to flatten
    :param coerce: apply coercion to top level, but not to sub ranges
    :return: flattened (coerced) items
    """
    if isinstance(data, collections.abc.Iterable) and not isinstance(
            data, (str, AddressRange, AddressCell)):
        for item in data:
            yield from flatten(item, coerce=coerce)
    else:
        yield coerce(data)


def uniqueify(seq):
    seen = set()
    return tuple(x for x in seq if x not in seen and not seen.add(x))


def is_number(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def coerce_to_number(value, convert_all=False):
    if value is None and convert_all:
        return 0

    if not isinstance(value, str):
        if isinstance(value, int):
            return int(value) if convert_all else value
        if is_number(value) and int(value) == float(value):
            return int(value)
        if is_array_arg(value):
            return coerce_to_number(value[0][0], convert_all)
        return value

    # True and False strings become numbers
    if convert_all and value.upper() in ('TRUE', 'FALSE', EMPTY):
        return int(len(value) == 4)

    try:
        if '.' not in value:
            return int(value)
    except (ValueError, TypeError):
        pass

    try:
        return float(value)
    except (ValueError, TypeError):
        return value


def coerce_to_string(value):
    if isinstance(value, bool):
        return str(value).upper()

    elif value is None:
        return ''

    elif not isinstance(value, str):
        return str(coerce_to_number(value))

    else:
        return value


def handle_ifs(args, op_range=None):
    """generic handler for ifs functions"""

    assert len(args) and len(args) % 2 == 0, \
        'Must have paired criteria and ranges'

    ranges = tuple(r if list_like(r) else ((r,),) for r in args[::2])

    # make sure all ranges are the same size
    sizes = {(len(a), len(a[0])) for a in ranges}
    if len(sizes) != 1:
        return VALUE_ERROR

    if op_range is not None:
        if not list_like(op_range):
            op_range = ((op_range, ), )

        size = len(op_range), len(op_range[0])
        for rng in ranges:  # pragma: no branch
            if size != (len(rng), len(rng[0])):
                return VALUE_ERROR

    # count the number of times a particular cell matches the criteria
    index_counts = collections.Counter(it.chain.from_iterable(
        find_corresponding_index(rng, criteria)
        for rng, criteria in zip(ranges, args[1::2])))

    ifs_count = len(args) // 2

    # if it is true in all cases, return the coordinates
    return tuple(idx for idx, cnt in index_counts.items() if cnt == ifs_count)


def build_wildcard_re(lookup_value):
    regex = QUESTION_MARK_RE.sub('.', STAR_RE.sub('.*', lookup_value))
    if regex != lookup_value:
        # this will be a regex match"""
        compiled = re.compile(f'^{regex.lower()}$')
        return lambda x: x is not None and compiled.match(x.lower()) is not None
    else:
        return None


def criteria_parser(criteria):
    """
    General rules:

        Criteria will be coerced to numbers,
        For equality comparisons, values will be coerced to numbers
        < and > will always be False when comparing strings to numbers
        <> will always be True when comparing strings to numbers

       You can use the wildcard characters—the question mark (?) and
       asterisk (*)—as the criteria argument. A question mark matches
       any single character; an asterisk matches any sequence of
       characters. If you want to find an actual question mark or
       asterisk, type a tilde (~) preceding the character.
    """

    if is_number(criteria):
        # numeric equals comparision
        criteria = coerce_to_number(criteria)

        def check(x):
            return is_number(x) and coerce_to_number(x) == criteria

    elif isinstance(criteria, str):
        match = OPERATORS_RE.match(criteria)
        criteria_operator = match.group('oper') or ''
        value = match.group('value')
        op = OPERATORS[criteria_operator]

        if op == operator.eq:

            if is_number(value):
                return criteria_parser(value)

            check = build_wildcard_re(value)
            if check is not None:
                return check

        if is_number(value):
            value = coerce_to_number(value)

            def check(x):
                if isinstance(x, str) or x is None:
                    # string always compare False unless '!='
                    return op == operator.ne
                else:
                    return op(x, value)
        else:
            value = value.lower()

            def check(x):
                """Compare with a string"""
                if x is None:
                    return (not value) != (op == operator.ne)

                elif not isinstance(x, str):
                    # non string always compare False unless '!='
                    return op == operator.ne
                else:
                    return op(x.lower(), value)

    else:
        raise ValueError(f"Couldn't parse criteria: {criteria}")

    return check


def find_corresponding_index(rng, criteria):
    return tuple(find_corresponding_index_generator(rng, criteria))


def find_corresponding_index_generator(rng, criteria):
    # parse criteria, build a criteria check
    check = criteria_parser(criteria)

    assert_list_like(rng)
    return ((r, c) for r, row in enumerate(rng)
            for c, item in enumerate(row) if check(item))


def list_like(data):
    return (not isinstance(data, (str, AddressRange, AddressCell)) and
            isinstance(data, collections.abc.Iterable))


def assert_list_like(data):
    if not list_like(data):
        raise TypeError(f'Must be a list like: {data}')


def type_cmp_value(value):
    """ Excel compares bools above strings which are above numbers

    https://stackoverflow.com/a/35051992/7311767

    :param value: Operand
    :return: tuple of type precedence and the default to use
    """
    if value in ERROR_CODES:
        return 3, value
    elif isinstance(value, bool):
        return 2, False
    elif isinstance(value, str):
        return 1, ''
    else:
        return 0, 0.0


class ExcelCmp(collections.namedtuple('ExcelCmp', 'cmp_type value empty')):

    def __new__(cls, value, empty=None):
        if isinstance(value, ExcelCmp):
            return value

        # empty as the searched for, becomes 0.0
        if value is None:
            cmp_type = 0 if empty is None else empty.cmp_type
            default_empty = 0.0 if empty is None else empty.empty
            value = default_empty
        else:
            cmp_type, default_empty = type_cmp_value(value)

        if cmp_type == 1:
            value = value.lower()

        return super(ExcelCmp, cls).__new__(cls, cmp_type, value, default_empty)

    def __lt__(self, other):
        other = ExcelCmp(other, empty=self)
        return super().__lt__(other)

    def __le__(self, other):
        other = ExcelCmp(other, empty=self)
        return super().__le__(other)

    def __gt__(self, other):
        other = ExcelCmp(other, empty=self)
        return super().__gt__(other)

    def __ge__(self, other):
        other = ExcelCmp(other, empty=self)
        return super().__ge__(other)

    def __eq__(self, other):
        other = ExcelCmp(other, empty=self)
        return self[0] == other[0] and self[1] == other[1]

    def __ne__(self, other):
        return not self == other


def build_operator_operand_fixup(capture_error_state):

    def array_fixup(left_op, op, right_op):
        """use numpy broadcasting for ranges"""
        # ::TODO:: this needs better error processing to match excel behavior
        left_op = np.array(left_op, dtype=object)
        right_op = np.array(right_op, dtype=object)
        b = np.broadcast(left_op, right_op)

        size = b.shape
        data = tuple(b)
        return tuple(
            tuple(fixup(u, op, v) for (u, v) in data[i: i + size[1]])
            for i in range(0, len(data), size[1])
        )

    def fixup(left_op, op, right_op):
        """Fix up python operations to be more excel like in these cases:

            Operand error

            Empty cells
            Case-insensitive string compare
            String to Number coercion
            String / Number multiplication
        """
        left_list, right_list = list_like(left_op), list_like(right_op)
        if not left_list and left_op in ERROR_CODES:
            return left_op

        if not right_list and right_op in ERROR_CODES:
            return right_op

        if left_list or right_list:
            return array_fixup(left_op, op, right_op)

        if op in COMPARISION_OPS:
            if left_op in (None, EMPTY):
                left_op = type_cmp_value(right_op)[1]

            if right_op in (None, EMPTY):
                right_op = type_cmp_value(left_op)[1]

            left_op = ExcelCmp(left_op)
            right_op = ExcelCmp(right_op)

        elif op == 'BitAnd':
            # use bitwise-and '&' as string concat not '+'
            op = 'Add'

            if left_op in (None, EMPTY):
                left_op = ''
            elif isinstance(left_op, bool):
                left_op = str(left_op).upper()
            elif isinstance(left_op, float) or isinstance(left_op, int):
                left_op = str(coerce_to_number(left_op))
            else:
                left_op = str(left_op)

            if right_op in (None, EMPTY):
                right_op = ''
            elif isinstance(right_op, bool):
                right_op = str(right_op).upper()
            elif isinstance(right_op, float) or isinstance(right_op, int):
                right_op = str(coerce_to_number(right_op))
            else:
                right_op = str(right_op)

        else:
            left_op = coerce_to_number(left_op, convert_all=True)
            right_op = coerce_to_number(right_op, convert_all=True)

            if not (is_number(left_op) and is_number(right_op) or
                    is_address(left_op) and is_address(right_op)):
                if op != 'USub':
                    capture_error_state(True, f'Values: {left_op} {op} {right_op}')
                    return VALUE_ERROR

        try:
            if op == 'USub':
                return PYTHON_AST_OPERATORS[op](right_op)
            else:
                return PYTHON_AST_OPERATORS[op](left_op, right_op)
        except ZeroDivisionError:
            capture_error_state(True, f'Values: {left_op} {op} {right_op}')
            return DIV0
        except TypeError:
            capture_error_state(True, f'Values: {left_op} {op} {right_op}')
            return VALUE_ERROR

    return fixup


class _IterativeEvalTracker:
    """When iteratively evaluating, keep track of which cycle we are on"""
    _ns = threading.local()

    @property
    def ns(self):
        if not hasattr(self._ns, 'todo'):
            self._ns.todo = set()
            self._ns.computed = set()
            self._ns.iteration_number = 0
        return self._ns

    def __call__(self, iterations=100, tolerance=0.001):
        self.ns.iteration_number = 0
        self.ns.iterations = iterations
        self.ns.tolerance = tolerance
        return self

    @property
    def tolerance(self):
        return self.ns.tolerance

    @property
    def done(self):
        return (self.ns.iteration_number >= self.ns.iterations or
                not self.ns.todo)

    def wip(self, cell):
        """Which cells are currently a Work In Progress"""
        self.ns.todo.add(cell)

    def calced(self, cell):
        """Mark which cells have been done this iteration"""
        self.ns.computed.add(cell)

    def is_calced(self, cell):
        """Which cells have been done this iteration"""
        return cell in self.ns.computed

    def inc_iteration_number(self):
        self.ns.iteration_number += 1
        self.ns.todo.clear()
        self.ns.computed.clear()


iterative_eval_tracker = _IterativeEvalTracker()
