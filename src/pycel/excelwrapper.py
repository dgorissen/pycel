# -*- coding: UTF-8 -*-
#
# Copyright 2011-2019 by Dirk Gorissen, Stephen Rauch and Contributors
# All rights reserved.
# This file is part of the Pycel Library, Licensed under GPLv3 (the 'License')
# You may not use this work except in compliance with the License.
# You may obtain a copy of the Licence at:
#   https://www.gnu.org/licenses/gpl-3.0.en.html

"""
    ExcelOpxWrapper : Can be run anywhere but only with post 2010 Excel formats
    ExcelOpxWrapperNoData  :
        Can be initialized with a instance of an OpenPyXl workbook
"""

import abc
import collections
import os
from unittest import mock

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import datetime as opxl_dt

from pycel.excelutil import AddressCell, AddressRange, flatten

ARRAY_FORMULA_NAME = '=CSE_INDEX'
ARRAY_FORMULA_FORMAT = '{}(%s,%s,%s,%s,%s)'.format(ARRAY_FORMULA_NAME)


class ExcelWrapper:
    __metaclass__ = abc.ABCMeta

    RangeData = collections.namedtuple('RangeData', 'address formula values')

    @abc.abstractmethod
    def get_range(self, address):
        """"""

    @abc.abstractmethod
    def get_used_range(self):
        """"""

    @abc.abstractmethod
    def get_active_sheet_name(self):
        """"""

    def get_formula_from_range(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)
        result = self.get_range(address)
        if isinstance(address, AddressCell):
            return result.formula if result.formula.startswith("=") else None
        else:
            return tuple(tuple(
                self.get_formula_from_range(a) for a in row
            ) for row in result.resolve_range)

    def get_formula_or_value(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)
        result = self.get_range(address)
        if isinstance(address, AddressCell):
            return result.formula or result.values
        else:
            return tuple(tuple(
                self.get_formula_or_value(a) for a in row
            ) for row in result.resolve_range)


class _OpxRange(ExcelWrapper.RangeData):
    """ Excel range wrapper that distributes reduced api used by compiler
        (Formula & Value)
    """
    def __new__(cls, cells, cells_dataonly, address):
        formula = None
        value = cells[0][0].value
        if isinstance(value, str) and value.startswith(ARRAY_FORMULA_NAME):
            # if this range refers to a CSE Array Formula, get the formula
            front, *args = cells[0][0].value[:-1].rsplit(',', 4)

            # if this range corresponds to the top left of a CSE Array formula
            if (args[0] == args[1] == '1') and all(
                    c.value and c.value.startswith(front)
                    for c in flatten(cells)):
                # apply formula to the range
                formula = '={%s}' % front[len(ARRAY_FORMULA_NAME) + 1:]
        else:
            formula = tuple(tuple(cls.cell_to_formula(cell) for cell in row)
                            for row in cells)

        values = tuple(tuple(cell.value for cell in row)
                       for row in cells_dataonly)
        return ExcelWrapper.RangeData.__new__(cls, address, formula, values)

    @classmethod
    def cell_to_formula(cls, cell):
        if cell.value is None:
            return ''
        else:
            formula = str(cell.value)
            if not formula.startswith('='):
                return ''

            elif formula.startswith('={') and formula[-1] == '}':
                # This is not in a CSE Array Context
                return '=index({},1,1)'.format(formula[1:])

            elif formula.startswith(ARRAY_FORMULA_NAME):
                # These are CSE Array formulas as encoded from sheet
                params = formula[len(ARRAY_FORMULA_NAME) + 1:-1].rsplit(',', 4)
                start_row = cell.row - int(params[1]) + 1
                start_col_idx = cell.col_idx - int(params[2]) + 1
                end_row = start_row + int(params[3]) - 1
                end_col_idx = start_col_idx + int(params[4]) - 1
                cse_range = AddressRange(
                    (start_col_idx, start_row, end_col_idx, end_row),
                    sheet=cell.parent.title)
                return '=index({},{},{})'.format(
                    cse_range.quoted_address, *params[1:3])
            else:
                return formula

    @property
    def resolve_range(self):
        return AddressRange(
            (self.address.start.col_idx,
             self.address.start.row,
             self.address.start.col_idx + len(self.values[0]) - 1,
             self.address.start.row + len(self.values) - 1),
            sheet=self.address.sheet
        ).resolve_range


class _OpxCell(_OpxRange):
    """ Excel cell wrapper that distributes reduced api used by compiler
        (Formula & Value)
    """
    def __new__(cls, cell, cell_dataonly, address):
        assert isinstance(address, AddressCell)
        return ExcelWrapper.RangeData.__new__(
            cls, address, cls.cell_to_formula(cell), cell_dataonly.value)


class ExcelOpxWrapper(ExcelWrapper):
    """ OpenPyXl implementation for ExcelWrapper interface """

    CfRule = collections.namedtuple(
        'CfRule', 'formula priority dxf_id dxf stop_if_true')

    def __init__(self, filename, app=None):
        super(ExcelWrapper, self).__init__()

        self.filename = os.path.abspath(filename)
        self._defined_names = None
        self._tables = None
        self._table_refs = {}
        self.workbook = None
        self.workbook_dataonly = None
        self._max_col_row = {}

    def max_col_row(self, sheet):
        if sheet not in self._max_col_row:
            worksheet = self.workbook[sheet]
            self._max_col_row[sheet] = worksheet.max_column, worksheet.max_row
        return self._max_col_row[sheet]

    @property
    def defined_names(self):
        if self.workbook is not None and self._defined_names is None:
            self._defined_names = {}

            for d_name in self.workbook.defined_names.definedName:
                destinations = [
                    (alias, wksht) for wksht, alias in d_name.destinations
                    if wksht in self.workbook]
                if len(destinations):
                    self._defined_names[str(d_name.name)] = destinations
        return self._defined_names

    def table(self, table_name):
        """ Return the table and the sheet it was found on

        :param table_name: name of table to retrieve
        :return: table, sheet_name
        """
        # table names are case insensitive
        if self._tables is None:
            TableAndSheet = collections.namedtuple(
                'TableAndSheet', 'table, sheet_name')
            self._tables = {
                t.name.lower(): TableAndSheet(t, ws.title)
                for ws in self.workbook for t in ws._tables}
            self._tables[None] = TableAndSheet(None, None)
        return self._tables.get(table_name.lower(), self._tables[None])

    def table_name_containing(self, address):
        """ Return the table name containing the address given """
        address = AddressCell(address)
        if address not in self._table_refs:
            for t in self.workbook[address.sheet]._tables:
                if address in AddressRange(t.ref):
                    self._table_refs[address] = t.name.lower()
                    break

        return self._table_refs.get(address)

    def conditional_format(self, address):
        """ Return the conditional formats applicable for this cell """
        address = AddressCell(address)
        all_formats = self.workbook[address.sheet].conditional_formatting
        formats = (cf for cf in all_formats if address.coordinate in cf)
        rules = []
        for cf in formats:
            origin = AddressRange(cf.cells.ranges[0].coord).start
            row_offset = address.row - origin.row
            col_offset = address.col_idx - origin.col_idx
            for rule in cf.rules:
                if rule.formula:
                    trans = Translator(
                        '={}'.format(rule.formula[0]), origin.coordinate)
                    formula = trans.translate_formula(
                        row_delta=row_offset, col_delta=col_offset)
                    rules.append(self.CfRule(
                        formula=formula,
                        priority=rule.priority,
                        dxf_id=rule.dxfId,
                        dxf=rule.dxf,
                        stop_if_true=rule.stopIfTrue,
                    ))
        return sorted(rules, key=lambda x: x.priority)

    def load(self):
        # work around type coercion to datetime that causes some issues
        with mock.patch('openpyxl.worksheet._reader.from_excel',
                        self.from_excel):
            self.workbook = load_workbook(self.filename)
            self.workbook_dataonly = load_workbook(
                self.filename, data_only=True)
        self.load_array_formulas()

    def load_array_formulas(self):
        # expand array formulas
        for ws in self.workbook:
            for address, props in ws.formula_attributes.items():
                if props.get('t') != 'array':
                    continue  # pragma: no cover

                # get the reference address for the array formula
                ref_addr = AddressRange(props.get('ref'))

                if isinstance(ref_addr, AddressRange):
                    formula = ws[address].value
                    for i, row in enumerate(ref_addr.rows, start=1):
                        for j, addr in enumerate(row, start=1):
                            ws[addr.coordinate] = ARRAY_FORMULA_FORMAT % (
                                formula[1:], i, j, *ref_addr.size)

    def set_sheet(self, s):
        self.workbook.active = self.workbook.index(self.workbook[s])
        self.workbook_dataonly.active = self.workbook_dataonly.index(
            self.workbook_dataonly[s])
        return self.workbook.active

    @staticmethod
    def from_excel(value, offset=opxl_dt.CALENDAR_WINDOWS_1900):
        # ::HACK:: excel thinks that 1900/02/29 was a thing.  In certain
        # circumstances openpyxl will return a datetime.  This is a problem
        # as we don't want them, and having been mapped to datetime
        # information may have been lost, so ignore the conversions
        return value

    def get_range(self, address):
        if not isinstance(address, (AddressRange, AddressCell)):
            address = AddressRange(address)

        if address.has_sheet:
            sheet = self.workbook[address.sheet]
            sheet_dataonly = self.workbook_dataonly[address.sheet]
        else:
            sheet = self.workbook.active
            sheet_dataonly = self.workbook_dataonly.active

        with mock.patch('openpyxl.worksheet._reader.from_excel',
                        self.from_excel):
            # work around type coercion to datetime that causes some issues

            if address.is_unbounded_range:
                # bound the address range to the data in the spreadsheet
                address = address & AddressRange(
                    (1, 1, *self.max_col_row(sheet.title)),
                    sheet=sheet.title)

            cells = sheet[address.coordinate]
            cells_dataonly = sheet_dataonly[address.coordinate]
            if isinstance(cells, (Cell, MergedCell)):
                return _OpxCell(cells, cells_dataonly, address)
            else:
                return _OpxRange(cells, cells_dataonly, address)

    def get_used_range(self):
        return self.workbook.active.iter_rows()

    def get_active_sheet_name(self):
        return self.workbook.active.title


class ExcelOpxWrapperNoData(ExcelOpxWrapper):
    """ ExcelWrapper interface from openpyxl workbook,
        without data_only workbook """

    @staticmethod
    def excel_value(formula, value):
        """A openpyxl sheet does not have values for formula cells"""
        return None if formula else value

    class OpxRange(_OpxRange):
        def __new__(cls, range_data):
            values = tuple(
                tuple(ExcelOpxWrapperNoData.excel_value(*cell)
                      for cell in zip(row_f, row_v))
                for row_f, row_v in zip(range_data.formula, range_data.values)
            )
            return ExcelWrapper.RangeData.__new__(
                cls, range_data.address, range_data.formula, values)

    class OpxCell(_OpxCell):
        def __new__(cls, cell_data):
            value = ExcelOpxWrapperNoData.excel_value(
                cell_data.formula, cell_data.values)
            return ExcelWrapper.RangeData.__new__(
                cls, cell_data.address, cell_data.formula, value)

    def __init__(self, workbook, filename='Unknown'):
        super().__init__(filename=filename)
        assert isinstance(workbook, Workbook)
        self.workbook = workbook
        self.workbook_dataonly = workbook
        self.load_array_formulas()

    def get_range(self, address):
        data = super().get_range(address)
        if isinstance(data.values, tuple):
            return self.OpxRange(data)
        else:
            return self.OpxCell(data)
